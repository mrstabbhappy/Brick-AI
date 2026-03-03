from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import Any, Dict, Iterable, Optional, Tuple

from dotenv import load_dotenv

                                             
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

                                                           
load_dotenv(os.path.join(_ROOT, ".env"), interpolate=False)

import openpyxl              

from lib.brick_pricing import upsert_brick_pricing              
from lib.mongo import get_db, get_db_name              
from lib.bu_locations_static import BU_LOCATIONS              


_ITEM_CODE_RE = re.compile(r"\(([A-Z]\d{4}A)\)")


def _norm(s: Any) -> str:
    return ("" if s is None else str(s)).strip()


def _norm_code(s: Any) -> str:
    return _norm(s).upper().replace(" ", "")


def _slugify_name(name: str) -> str:
    v = (name or "").strip().lower()
    v = re.sub(r"\s+", "-", v)
    v = re.sub(r"[^a-z0-9\-]", "", v)
    v = re.sub(r"\-+", "-", v).strip("-")
    return v


def _strip_tw_prefix(region: str) -> str:
    v = (region or "").strip()
    if not v:
        return v
                                                                         
    v2 = re.sub(r"^\s*taylor\s+wimpey\s+", "", v, flags=re.IGNORECASE)
    return v2.strip()


def _canonicalize_region_name(region: str) -> str:
    v = (region or "").strip()
    if not v:
        return v
    key = v.lower()
    aliases = {
        "north midland": "North Midlands",
        "southern": "Southern Counties",
    }
    return aliases.get(key, v)


def _bu_name_to_id_map() -> Dict[str, str]:
    out: Dict[str, str] = {}
    for loc in BU_LOCATIONS:
        bu_id = (loc.get("id") or "").strip()
        name = (loc.get("name") or "").strip()
        if bu_id:
            out[bu_id.lower()] = bu_id
        if name:
            out[name.lower()] = bu_id
            out[_slugify_name(name)] = bu_id
    return out


def _match_bu_id(
    raw_region: str,
    *,
    bu_map: Dict[str, str],
    allow_unknown: bool = False,
) -> Optional[str]:
    v = _canonicalize_region_name(_strip_tw_prefix(_norm(raw_region)))
    if not v:
        return None
    key = v.lower()
    if key in bu_map:
        return bu_map[key]
    slug = _slugify_name(v)
    if slug in bu_map:
        return bu_map[slug]
    return slug if (allow_unknown and slug) else None


@dataclass(frozen=True)
class FactoryColumn:
    col_idx: int
    header_name: str
    factory_code: str


def _find_header_row(ws) -> int:
    for r in range(1, 60):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 60)]
        if any(isinstance(v, str) and "2nd Item Number" in v for v in vals if v is not None):
            return r
    raise RuntimeError("Could not find header row containing '2nd Item Number'")


def _parse_factory_columns(ws, header_row: int) -> list[FactoryColumn]:
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, 200)]
                                                                                      
    try:
        t4_idx = headers.index("T4") + 1           
    except ValueError:
        raise RuntimeError("Header row missing 'T4'")

    code_row = header_row + 1                                                      
    out: list[FactoryColumn] = []

    for c in range(t4_idx + 1, len(headers) + 1):
        h = headers[c - 1]
        if not h:
            continue
        header_name = _norm(h)
        code = _norm_code(ws.cell(row=code_row, column=c).value)
                                                                    
        if not code:
            continue
        if len(code) < 2:
            continue
        out.append(FactoryColumn(col_idx=c, header_name=header_name, factory_code=code))

    if not out:
        raise RuntimeError("No factory/site columns detected (after T4)")

    return out


def _iter_data_rows(ws, start_row: int) -> Iterable[int]:
    r = start_row
    while True:
        item = _norm(ws.cell(row=r, column=2).value)
        if not item:
            break
        yield r
        r += 1


def _parse_tier_prices(ws, row: int) -> Dict[str, Any]:
                              
    out: Dict[str, Any] = {}
    for band, col in (("T1", 5), ("T2", 6), ("T3", 7), ("T4", 8)):
        v = ws.cell(row=row, column=col).value
        if v is None or _norm(v) == "":
            continue
        out[band] = v
    return out


def _parse_region_raw(ws, row: int) -> str:
                                                                                         
    v11 = _norm(ws.cell(row=row, column=11).value)
    v10 = _norm(ws.cell(row=row, column=10).value)
    return v11 or v10


def _extract_codes_from_text(text: str) -> list[str]:
    t = _norm(text)
    if not t:
        return []
    m = _ITEM_CODE_RE.search(t)
    if m:
        return [_norm_code(m.group(1))]
    return []


def _build_itemcode_to_brickid_map(db) -> Tuple[Dict[str, str], Dict[str, list[str]]]:
    bricks = db.bricks
    mapping: Dict[str, str] = {}
    dups: Dict[str, list[str]] = defaultdict(list)

    cursor = bricks.find(
        {"dataset": "catalog"},
        {
            "brick_id": 1,
            "_id": 1,
            "metadata.item_number": 1,
            "metadata.item_code": 1,
            "metadata.display_name": 1,
            "metadata.brick_name": 1,
        },
    )
    for d in cursor:
        brick_id = str(d.get("brick_id") or d.get("_id") or "").strip()
        if not brick_id:
            continue

        md = d.get("metadata") or {}
        candidates: list[str] = []
        for key in ("item_number", "item_code"):
            code = _norm_code(md.get(key))
            if code:
                candidates.append(code)
        candidates += _extract_codes_from_text(md.get("display_name") or "")
        candidates += _extract_codes_from_text(md.get("brick_name") or "")

        for code in dict.fromkeys(candidates):
            if code in mapping and mapping[code] != brick_id:
                dups[code].append(brick_id)
                continue
            mapping[code] = brick_id

    return mapping, dups


def _upsert_bu_factory_pricing_rules(
    db,
    tier_votes: Dict[Tuple[str, str], Counter],
    *,
    dry_run: bool,
) -> dict:
    coll = db.bu_factory_pricing
    inserted = 0
    updated = 0
    conflicts = 0

    for (bu_code, factory_code), counter in tier_votes.items():
        if not bu_code or not factory_code or not counter:
            continue
        (tier, count) = counter.most_common(1)[0]
        if len(counter) > 1:
            conflicts += 1

        doc = {
            "bu_code": bu_code,
            "factory_code": factory_code,
            "price_band": tier,
            "effective_to": None,
            "source": "SiteMatrix.xlsx",
            "votes": dict(counter),
        }

        if dry_run:
            continue

        res = coll.update_one(
            {
                "bu_code": bu_code,
                "factory_code": factory_code,
                "$or": [{"effective_to": None}, {"effective_to": {"$exists": False}}],
            },
            {"$set": doc},
            upsert=True,
        )
        if res.upserted_id is not None:
            inserted += 1
        elif res.modified_count:
            updated += 1

    return {"inserted": inserted, "updated": updated, "conflicts": conflicts, "total": len(tier_votes)}


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(
        description="Import tier ASPs + BU/factory tier rules from tools/SiteMatrix.xlsx into Mongo"
    )
    p.add_argument(
        "--file",
        default=os.path.join(_ROOT, "tools", "SiteMatrix.xlsx"),
        help="Path to SiteMatrix.xlsx",
    )
    p.add_argument("--dry-run", action="store_true", help="Parse and report, but do not write to Mongo")
    p.add_argument(
        "--write-bu-factory-pricing",
        action="store_true",
        help="Also seed bu_factory_pricing (mode tier per BU+factory from the matrix)",
    )
    p.add_argument(
        "--allow-unknown-bus",
        action="store_true",
        help="Allow BU codes not in lib/bu_locations_static.py (fallback to slugified region name)",
    )
    p.add_argument(
        "--limit-rows",
        type=int,
        default=0,
        help="Optional safety limit (0 = no limit)",
    )
    args = p.parse_args(argv)

    db = get_db()
    print(f"DB: {get_db_name()}")

    path = args.file
    if not os.path.exists(path):
        raise SystemExit(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    factory_cols = _parse_factory_columns(ws, header_row)

    item_to_prices: Dict[str, Dict[str, Any]] = {}
    item_to_home_factory: Dict[str, str] = {}
    tier_votes: Dict[Tuple[str, str], Counter] = defaultdict(Counter)

    bu_map = _bu_name_to_id_map()

    processed_rows = 0
    unmatched_regions = Counter()

    start_row = header_row + 2                                      

    for r in _iter_data_rows(ws, start_row):
        processed_rows += 1
        if args.limit_rows and processed_rows > args.limit_rows:
            break

        item_code = _norm_code(ws.cell(row=r, column=2).value)
        if not item_code:
            continue

        home_factory_code = _norm_code(ws.cell(row=r, column=1).value)
        if home_factory_code and item_code not in item_to_home_factory:
            item_to_home_factory[item_code] = home_factory_code

        prices = _parse_tier_prices(ws, r)
        if prices:
            prev = item_to_prices.get(item_code)
            if prev is None:
                item_to_prices[item_code] = prices
            else:
                                                                        
                if any(prev.get(k) != prices.get(k) for k in ("T1", "T2", "T3", "T4")):
                                             
                    item_to_prices[item_code] = prev

        if args.write_bu_factory_pricing:
            raw_region = _parse_region_raw(ws, r)
            bu_code = _match_bu_id(raw_region, bu_map=bu_map, allow_unknown=bool(args.allow_unknown_bus))
            if not bu_code:
                unmatched_regions[raw_region or "(blank)"] += 1
                continue

            for fc in factory_cols:
                tier = _norm(ws.cell(row=r, column=fc.col_idx).value)
                if tier not in ("T1", "T2", "T3", "T4"):
                    continue
                tier_votes[(bu_code, fc.factory_code)][tier] += 1

                                       
    item_to_brick_id, duplicates = _build_itemcode_to_brickid_map(db)

                          
    updated_pricing = 0
    missing_bricks = 0
    backfilled_brick_factory = 0
    backfill_factory_conflicts = 0

    bricks_coll = db.bricks

    for item_code, prices in item_to_prices.items():
        brick_id = item_to_brick_id.get(item_code)
        if not brick_id:
            missing_bricks += 1
            continue

        home_factory_code = item_to_home_factory.get(item_code) or None

        band_asps = {f"asp_{k}": prices.get(k) for k in ("T1", "T2", "T3", "T4") if k in prices}
        if not band_asps:
            continue

        if not args.dry_run:
            upsert_brick_pricing(brick_id=brick_id, band_asps=band_asps, factory_code=home_factory_code)

                                                                                                
                                                                                 
            if home_factory_code:
                b = bricks_coll.find_one(
                    {"brick_id": brick_id},
                    {"factory_code": 1, "metadata.factory_code": 1},
                )
                existing_root = _norm_code((b or {}).get("factory_code"))
                existing_md = _norm_code(((b or {}).get("metadata") or {}).get("factory_code"))
                existing = existing_root or existing_md
                if not existing:
                    bricks_coll.update_one(
                        {"brick_id": brick_id},
                        {"$set": {"factory_code": home_factory_code, "metadata.factory_code": home_factory_code}},
                    )
                    backfilled_brick_factory += 1
                elif existing != home_factory_code:
                    backfill_factory_conflicts += 1
        updated_pricing += 1

    bu_rules_summary = None
    if args.write_bu_factory_pricing:
        bu_rules_summary = _upsert_bu_factory_pricing_rules(db, tier_votes, dry_run=args.dry_run)

    print("--- Summary ---")
    print({
        "rows_processed": processed_rows,
        "unique_item_codes": len(item_to_prices),
        "pricing_upserts": updated_pricing,
        "pricing_missing_brick_match": missing_bricks,
        "bricks_factory_backfilled": backfilled_brick_factory,
        "bricks_factory_conflicts": backfill_factory_conflicts,
        "duplicate_item_codes_in_db": len(duplicates),
        "unmatched_regions": sum(unmatched_regions.values()),
        "bu_factory_pricing": bu_rules_summary,
        "dry_run": bool(args.dry_run),
    })

    if duplicates:
        sample = dict(list(duplicates.items())[:10])
        print("Duplicate item codes found in catalog (showing up to 10):")
        print(sample)

    if unmatched_regions:
        sample = unmatched_regions.most_common(10)
        print("Unmatched regions (showing up to 10):")
        print(sample)

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
